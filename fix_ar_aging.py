"""
fix_ar_aging.py — Refresh CashFlowCompass AR/AP workbook formulas and aging logic.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


WORKBOOK_NAME = "SignalStack_CashFlowCompass.xlsx"
ROOT_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT_DIR / WORKBOOK_NAME


AR_DATA = [
    ("Acme Petroleum", "INV-1042", date(2026, 3, 10), date(2026, 4, 9), 4850),
    ("Summit Logistics", "INV-1045", date(2026, 3, 18), date(2026, 4, 17), 2200),
    ("High Plains Co", "INV-1038", date(2026, 2, 15), date(2026, 3, 17), 6100),
    ("Front Range Inc", "INV-1035", date(2026, 2, 1), date(2026, 3, 3), 1875),
    ("Mountain Fuel", "INV-1030", date(2026, 1, 10), date(2026, 2, 9), 3300),
    ("Valley Services", "INV-1025", date(2025, 12, 15), date(2026, 1, 14), 2750),
    ("Peak Operations", "INV-1020", date(2025, 11, 1), date(2025, 12, 1), 5200),
]

AP_DATA = [
    ("Fuel Supplier Co", "AP-2301", date(2026, 3, 1), date(2026, 3, 25), 8400),
    ("Fleet Maintenance", "AP-2305", date(2026, 3, 10), date(2026, 4, 9), 2200),
    ("Software Stack", "AP-2310", date(2026, 3, 15), date(2026, 4, 14), 890),
    ("Office Supplies", "AP-2315", date(2026, 3, 20), date(2026, 4, 19), 340),
    ("Insurance Premium", "AP-2320", date(2026, 3, 25), date(2026, 4, 9), 3100),
    ("Equipment Lease", "AP-2325", date(2026, 3, 28), date(2026, 4, 28), 1850),
]


def apply_ar_aging_updates(ws) -> None:
    """Update AR Aging detail rows, formulas, summary formulas, and conditional formatting."""
    for idx, (customer, invoice_no, invoice_date, due_date, amount) in enumerate(AR_DATA, start=4):
        ws[f"A{idx}"] = customer
        ws[f"B{idx}"] = invoice_no
        ws[f"C{idx}"] = invoice_date
        ws[f"D{idx}"] = due_date
        ws[f"E{idx}"] = amount
        ws[f"F{idx}"] = f"=TODAY()-C{idx}"
        ws[f"G{idx}"] = (
            f'=IF(F{idx}<=30,"Current (0-30d)",'
            f'IF(F{idx}<=60,"31-60 days",'
            f'IF(F{idx}<=90,"61-90 days",'
            f'IF(F{idx}<=120,"91-120 days","120+ days"))))'
        )

        ws[f"C{idx}"].number_format = "mm/dd/yyyy"
        ws[f"D{idx}"].number_format = "mm/dd/yyyy"
        ws[f"E{idx}"].number_format = "$#,##0"
        ws[f"F{idx}"].number_format = "0"

    ws["A12"] = "AGING SUMMARY"
    ws["A13"] = "Current (0-30d)"
    ws["A14"] = "31-60 days"
    ws["A15"] = "61-90 days"
    ws["A16"] = "91-120 days"
    ws["A17"] = "120+ days"
    ws["A18"] = "Total Outstanding"

    ws["B13"] = "=SUMPRODUCT((F4:F10<=30)*E4:E10)"
    ws["B14"] = "=SUMPRODUCT((F4:F10>30)*(F4:F10<=60)*E4:E10)"
    ws["B15"] = "=SUMPRODUCT((F4:F10>60)*(F4:F10<=90)*E4:E10)"
    ws["B16"] = "=SUMPRODUCT((F4:F10>90)*(F4:F10<=120)*E4:E10)"
    ws["B17"] = "=SUMPRODUCT((F4:F10>120)*E4:E10)"
    ws["B18"] = "=SUM(E4:E10)"

    for idx in range(13, 19):
        ws[f"B{idx}"].number_format = "$#,##0"

    # Rebuild AR conditional formatting so repeated script runs do not duplicate rules.
    ws.conditional_formatting._cf_rules.clear()

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    ws.conditional_formatting.add(
        "G4:G10",
        FormulaRule(formula=["=AND($F4>=0,$F4<=30)"], stopIfTrue=True, fill=green_fill),
    )
    ws.conditional_formatting.add(
        "G4:G10",
        FormulaRule(formula=["=AND($F4>30,$F4<=60)"], stopIfTrue=True, fill=light_yellow_fill),
    )
    ws.conditional_formatting.add(
        "G4:G10",
        FormulaRule(formula=["=AND($F4>60,$F4<=90)"], stopIfTrue=True, fill=yellow_fill),
    )
    ws.conditional_formatting.add(
        "G4:G10",
        FormulaRule(formula=["=AND($F4>90,$F4<=120)"], stopIfTrue=True, fill=orange_fill),
    )
    ws.conditional_formatting.add(
        "G4:G10",
        FormulaRule(formula=["=$F4>120"], stopIfTrue=True, fill=red_fill),
    )


def apply_ap_schedule_updates(ws) -> None:
    """Update AP Schedule detail rows, formulas, and formatting."""
    for idx, (vendor, invoice_no, invoice_date, due_date, amount) in enumerate(AP_DATA, start=4):
        ws[f"A{idx}"] = vendor
        ws[f"B{idx}"] = invoice_no
        ws[f"C{idx}"] = invoice_date
        ws[f"D{idx}"] = due_date
        ws[f"E{idx}"] = amount
        ws[f"F{idx}"] = f"=D{idx}-TODAY()"
        ws[f"G{idx}"] = (
            f'=IF(F{idx}<0,"OVERDUE",'
            f'IF(F{idx}<=7,"DUE THIS WEEK",'
            f'IF(F{idx}<=14,"DUE SOON","SCHEDULED")))'
        )

        ws[f"C{idx}"].number_format = "mm/dd/yyyy"
        ws[f"D{idx}"].number_format = "mm/dd/yyyy"
        ws[f"E{idx}"].number_format = "$#,##0"
        ws[f"F{idx}"].number_format = "0"


def apply_weekly_position_links(ws) -> None:
    """Link Weekly Position AR/AP summary cells to AR Aging and AP Schedule tabs."""
    ws["B6"] = "=SUMPRODUCT((('AR Aging'!F4:F10<=30)*'AR Aging'!E4:E10))"
    ws["C6"] = "=SUMPRODUCT((ABS('AP Schedule'!F4:F9)<=30)*'AP Schedule'!E4:E9)"


def find_delivery_excel_folder(root: Path) -> Path | None:
    """Find a likely delivery folder at project root level, if present."""
    candidates = [
        root / "delivery_excel",
        root / "delivery",
        root / "deliverables",
        root / "reports" / "delivery_excel",
        root / "reports" / "delivery",
        root / "reports" / "deliverables",
    ]
    for folder in candidates:
        if folder.is_dir():
            return folder
    return None


def verify_changes(path: Path) -> bool:
    """Reload workbook and verify key formulas were persisted."""
    wb = load_workbook(path, data_only=False)
    ws_ar = wb["AR Aging"]
    ws_ap = wb["AP Schedule"]
    ws_weekly = wb["Weekly Position"]

    checks = {
        "AR days formula F4": ws_ar["F4"].value == "=TODAY()-C4",
        "AR bucket formula G4": isinstance(ws_ar["G4"].value, str) and "Current (0-30d)" in ws_ar["G4"].value,
        "AR summary B17": ws_ar["B17"].value == "=SUMPRODUCT((F4:F10>120)*E4:E10)",
        "AR summary B18": ws_ar["B18"].value == "=SUM(E4:E10)",
        "AP days formula F4": ws_ap["F4"].value == "=D4-TODAY()",
        "AP priority formula G4": isinstance(ws_ap["G4"].value, str) and "DUE THIS WEEK" in ws_ap["G4"].value,
        "Weekly B6 link": ws_weekly["B6"].value == "=SUMPRODUCT((('AR Aging'!F4:F10<=30)*'AR Aging'!E4:E10))",
        "Weekly C6 link": ws_weekly["C6"].value == "=SUMPRODUCT((ABS('AP Schedule'!F4:F9)<=30)*'AP Schedule'!E4:E9)",
    }

    all_ok = True
    for label, passed in checks.items():
        status = "OK" if passed else "FAIL"
        print(f"[fix_ar_aging] Verify {label}: {status}")
        all_ok = all_ok and passed
    return all_ok


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"[fix_ar_aging] Workbook not found: {WORKBOOK_PATH}")

    print(f"[fix_ar_aging] Loading workbook: {WORKBOOK_PATH}")
    wb = load_workbook(WORKBOOK_PATH, data_only=False)

    ws_ar = wb["AR Aging"]
    ws_ap = wb["AP Schedule"]
    ws_weekly = wb["Weekly Position"]

    apply_ar_aging_updates(ws_ar)
    print("[fix_ar_aging] AR Aging tab updated: detail rows, formulas, 5-tier summary, conditional formatting.")

    apply_ap_schedule_updates(ws_ap)
    print("[fix_ar_aging] AP Schedule tab updated: detail rows and live due-priority formulas.")

    apply_weekly_position_links(ws_weekly)
    print("[fix_ar_aging] Weekly Position links updated: B6 and C6 are now dynamic formulas.")

    wb.save(WORKBOOK_PATH)
    print(f"[fix_ar_aging] Saved workbook: {WORKBOOK_PATH}")

    delivery_folder = find_delivery_excel_folder(ROOT_DIR)
    if delivery_folder is not None:
        delivery_path = delivery_folder / WORKBOOK_NAME
        copy2(WORKBOOK_PATH, delivery_path)
        print(f"[fix_ar_aging] Saved delivery copy: {delivery_path}")
    else:
        print("[fix_ar_aging] Delivery Excel folder not found at expected locations; skipped copy.")

    verified = verify_changes(WORKBOOK_PATH)
    if verified:
        print("[fix_ar_aging] Verification complete: all targeted formulas are in place.")
    else:
        print("[fix_ar_aging] Verification finished with failures. Inspect workbook formulas.")


if __name__ == "__main__":
    main()
