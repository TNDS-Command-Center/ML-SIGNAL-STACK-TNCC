"""
fix_root_workbooks.py
=====================
Normalize root workbook input tabs so date/number fields are stored as real
Excel types (not text). This avoids "Text to Columns" fixes reverting.

Usage:
    python fix_root_workbooks.py
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import openpyxl


BASE_DIR = Path(__file__).resolve().parent


WORKBOOK_RULES = {
    "SignalStack_PipelinePulse.xlsx": {
        "Pipeline Log": {
            "header_row": 3,
            "start_row": 4,
            "columns": {
                "Date Entered": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Last Activity": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Next Action Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Est. Value": {"kind": "number", "numfmt": "$#,##0"},
            },
        }
    },
    "SignalStack_OpsPulse.xlsx": {
        "Weekly Log": {
            "header_row": 3,
            "start_row": 5,
            "columns": {
                "Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Jobs\nCompleted": {"kind": "number", "numfmt": "0"},
                "Scheduled\nTime (h)": {"kind": "number", "numfmt": "0.0"},
                "Actual\nTime (h)": {"kind": "number", "numfmt": "0.0"},
            },
        }
    },
    "SignalStack_TeamTempo.xlsx": {
        "Hours Log": {
            "header_row": 3,
            "start_row": 4,
            "format_through_row": 2000,
            "columns": {
                "Week Of": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Regular Hrs": {"kind": "number", "numfmt": "0.0"},
                "OT Hrs": {"kind": "number", "numfmt": "0.0"},
                "Training Hrs": {"kind": "number", "numfmt": "0.0"},
                "PTO Hrs": {"kind": "number", "numfmt": "0.0"},
                "Total Hrs": {"kind": "number", "numfmt": "0.0"},
            },
        },
        "Roster": {
            "header_row": 3,
            "start_row": 4,
            "format_through_row": 500,
            "columns": {
                "Start Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Training Due": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Pay Rate": {"kind": "number", "numfmt": "$#,##0.00"},
            },
        },
    },
    "tnds-sales-data-template.xlsx": {
        "RAW_INPUT": {
            "header_row": 3,
            "start_row": 5,
            "columns": {
                "Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Qty": {"kind": "number", "numfmt": "0.00"},
                "Sales Price": {"kind": "number", "numfmt": "$#,##0.00"},
                "Tax": {"kind": "number", "numfmt": "$#,##0.00"},
            },
        }
    },
    "SignalStack_CashFlowCompass.xlsx": {
        "Weekly Position": {
            "header_row": 9,
            "start_row": 10,
            "columns": {
                "Week Of": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Cash on Hand": {"kind": "number", "numfmt": "$#,##0"},
                "AR Collected": {"kind": "number", "numfmt": "$#,##0"},
                "AP Paid": {"kind": "number", "numfmt": "$#,##0"},
                "Revenue In": {"kind": "number", "numfmt": "$#,##0"},
                "Expenses Out": {"kind": "number", "numfmt": "$#,##0"},
            },
        },
        "AR Aging": {
            "header_row": 3,
            "start_row": 4,
            "columns": {
                "Invoice Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Due Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Amount": {"kind": "number", "numfmt": "$#,##0"},
            },
        },
        "AP Schedule": {
            "header_row": 3,
            "start_row": 4,
            "columns": {
                "Invoice Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Due Date": {"kind": "date", "numfmt": "mm/dd/yyyy"},
                "Amount": {"kind": "number", "numfmt": "$#,##0"},
            },
        },
    },
}


def parse_date_value(value) -> Optional[datetime]:
    """Convert a date-like value into datetime if possible."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None

    formats = (
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    )
    for fmt in formats:
        try:
            dt = datetime.strptime(text, fmt)
            return datetime(dt.year, dt.month, dt.day)
        except ValueError:
            continue
    return None


def parse_number_value(value) -> Optional[float]:
    """Convert a number-like value into int/float when possible."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return None

    text = value.strip()
    if not text:
        return None
    if text.startswith("="):
        return None

    clean = text.replace("$", "").replace(",", "").strip()
    if clean.endswith("%"):
        clean = clean[:-1].strip()
        try:
            return float(clean) / 100.0
        except ValueError:
            return None
    try:
        num = float(clean)
    except ValueError:
        return None

    return int(num) if num.is_integer() else num


def header_map(ws, header_row: int) -> Dict[str, int]:
    """Map header text to column index for a worksheet."""
    mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if isinstance(value, str) and value.strip():
            mapping[value.strip()] = col_idx
    return mapping


def normalize_sheet(ws, rules: dict) -> dict:
    """Normalize configured columns in a sheet and return counters."""
    hdr = header_map(ws, rules["header_row"])
    start_row = rules["start_row"]
    format_through_row = rules.get("format_through_row")
    counters = {
        "date_converted": 0,
        "number_converted": 0,
        "formats_applied": 0,
        "template_formats_applied": 0,
        "rows_touched": 0,
    }

    for col_name, spec in rules["columns"].items():
        col_idx = hdr.get(col_name)
        if not col_idx:
            print(f"[root_fix] WARNING: '{ws.title}' missing column '{col_name}'")
            continue

        row_touched = set()
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(r, col_idx)
            value = cell.value

            if isinstance(value, str) and value.startswith("="):
                continue
            if value is None or (isinstance(value, str) and not value.strip()):
                continue

            if spec["kind"] == "date":
                parsed = parse_date_value(value)
                if parsed is not None:
                    if not isinstance(value, (datetime, date)):
                        cell.value = parsed
                        counters["date_converted"] += 1
                        row_touched.add(r)
                    if cell.number_format != spec["numfmt"]:
                        cell.number_format = spec["numfmt"]
                        counters["formats_applied"] += 1
                        row_touched.add(r)

            elif spec["kind"] == "number":
                parsed = parse_number_value(value)
                if parsed is not None:
                    if not isinstance(value, (int, float)):
                        cell.value = parsed
                        counters["number_converted"] += 1
                        row_touched.add(r)
                    if cell.number_format != spec["numfmt"]:
                        cell.number_format = spec["numfmt"]
                        counters["formats_applied"] += 1
                        row_touched.add(r)

        counters["rows_touched"] += len(row_touched)

        # Template formatting: apply number formats to future entry rows so new
        # pasted values inherit the intended date/number format.
        if format_through_row and format_through_row >= start_row:
            for r in range(start_row, format_through_row + 1):
                cell = ws.cell(r, col_idx)
                if cell.number_format != spec["numfmt"]:
                    cell.number_format = spec["numfmt"]
                    counters["template_formats_applied"] += 1

    return counters


def sanitize_formula_prefixes(ws) -> int:
    """
    Remove Excel compatibility prefixes that can break formulas in some installs
    after round-tripping through editors.
    """
    sanitized = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            value = cell.value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            new_value = value.replace("_xlfn.", "").replace("_xludf.", "")
            if new_value != value:
                cell.value = new_value
                sanitized += 1
    return sanitized


def repair_team_tempo_dashboard(ws) -> dict:
    """
    Repair known Team Tempo dashboard formula issues:
    1) Replace _xludf.ISOWEEKNUM with ISOWEEKNUM for Excel compatibility.
    2) Restore weekly Headcount formula in column B trend grid.
    """
    week_formula_written = 0
    headcount_written = 0

    # Column A ISO week label formulas.
    # Use WEEKNUM(...,21) for compatibility across Excel installs.
    for r in range(10, 2001):
        formula = (
            f'=IF($J{r}="","",YEAR($J{r}-WEEKDAY($J{r},2)+4)&"-W"&'
            f'TEXT(WEEKNUM($J{r},21),"00"))'
        )
        cell = ws.cell(r, 1)
        if cell.value != formula:
            cell.value = formula
            week_formula_written += 1

    # Restore headcount trend formula in column B.
    # Uses active roster entries started on/before each week start.
    for r in range(10, 2001):
        formula = (
            f'=IF($J{r}="","",COUNTIFS('
            f'Roster!$D$4:$D$500,"Active",'
            f'Roster!$C$4:$C$500,"<="&$J{r}))'
        )
        cell = ws.cell(r, 2)
        if cell.value != formula:
            cell.value = formula
            headcount_written += 1
        if cell.number_format != "0":
            cell.number_format = "0"

    return {
        "week_formula_written": week_formula_written,
        "headcount_written": headcount_written,
    }


def repair_ops_dashboard(ws) -> dict:
    """
    Repair Ops dashboard ISO week formulas for broad Excel compatibility.
    """
    week_formula_written = 0
    for r in range(10, 2001):
        formula = (
            f'=IF($J{r}="","",YEAR($J{r}-WEEKDAY($J{r},2)+4)&"-W"&'
            f'TEXT(WEEKNUM($J{r},21),"00"))'
        )
        cell = ws.cell(r, 1)
        if cell.value != formula:
            cell.value = formula
            week_formula_written += 1
    return {"week_formula_written": week_formula_written}


def repair_pipeline_dashboard(ws) -> dict:
    """
    Repair Pipeline dashboard date formula compatibility.
    Replaces MINIFS with AGGREGATE-based smallest-date selection.
    """
    updated = 0
    cell = ws["G16"]
    formula = (
        '=IFERROR(AGGREGATE(15,6,'
        "'Pipeline Log'!J:J/((\'Pipeline Log\'!J:J>=TODAY())*"
        "(\'Pipeline Log\'!D:D<>\"Closed Lost\")),1),\"\")"
    )
    if cell.value != formula:
        cell.value = formula
        updated += 1
    if cell.number_format != "mm/dd/yyyy":
        cell.number_format = "mm/dd/yyyy"
    return {"next_follow_up_formula_written": updated}


def run_root_workbook_fixes() -> dict:
    """Run root workbook normalization and return summary counters."""
    print("[root_fix] Scanning root workbooks...")
    changed_files = 0
    errors = []

    for workbook_name, sheet_rules in WORKBOOK_RULES.items():
        path = BASE_DIR / workbook_name
        if not path.exists():
            print(f"[root_fix] SKIP: missing {workbook_name}")
            continue

        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            file_changed = False
            print(f"[root_fix] Processing {workbook_name}")

            for sheet_name, rules in sheet_rules.items():
                if sheet_name not in wb.sheetnames:
                    print(f"[root_fix] WARNING: missing sheet '{sheet_name}'")
                    continue
                ws = wb[sheet_name]
                result = normalize_sheet(ws, rules)
                touched = (
                    result["date_converted"]
                    + result["number_converted"]
                    + result["formats_applied"]
                    + result["template_formats_applied"]
                )
                if touched > 0:
                    file_changed = True
                print(
                    f"[root_fix] {sheet_name}: date_converted={result['date_converted']}, "
                    f"number_converted={result['number_converted']}, "
                    f"formats_applied={result['formats_applied']}, "
                    f"template_formats_applied={result['template_formats_applied']}, "
                    f"rows_touched={result['rows_touched']}"
                )

            # Formula compatibility cleanup across all sheets.
            formula_sanitized = 0
            for ws in wb.worksheets:
                formula_sanitized += sanitize_formula_prefixes(ws)
            if formula_sanitized > 0:
                file_changed = True
            print(f"[root_fix] Formula prefix sanitizations: {formula_sanitized}")

            if workbook_name == "SignalStack_OpsPulse.xlsx" and "Dashboard" in wb.sheetnames:
                dash = repair_ops_dashboard(wb["Dashboard"])
                if dash["week_formula_written"] > 0:
                    file_changed = True
                print(
                    f"[root_fix] Dashboard: week_formula_written={dash['week_formula_written']}"
                )
                wb.calculation.calcMode = "auto"
                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True

            if workbook_name == "SignalStack_TeamTempo.xlsx" and "Dashboard" in wb.sheetnames:
                dash = repair_team_tempo_dashboard(wb["Dashboard"])
                dash_touched = dash["week_formula_written"] + dash["headcount_written"]
                if dash_touched > 0:
                    file_changed = True
                print(
                    f"[root_fix] Dashboard: week_formula_written={dash['week_formula_written']}, "
                    f"headcount_written={dash['headcount_written']}"
                )

                # Force automatic recalc on open.
                wb.calculation.calcMode = "auto"
                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True

            if workbook_name == "SignalStack_PipelinePulse.xlsx" and "Dashboard" in wb.sheetnames:
                dash = repair_pipeline_dashboard(wb["Dashboard"])
                if dash["next_follow_up_formula_written"] > 0:
                    file_changed = True
                print(
                    "[root_fix] Dashboard: "
                    f"next_follow_up_formula_written={dash['next_follow_up_formula_written']}"
                )
                wb.calculation.calcMode = "auto"
                wb.calculation.fullCalcOnLoad = True
                wb.calculation.forceFullCalc = True

            if file_changed:
                wb.save(path)
                changed_files += 1
                print(f"[root_fix] SAVED: {workbook_name}")
            else:
                print(f"[root_fix] No changes needed: {workbook_name}")
        except Exception as e:
            msg = f"{workbook_name}: {e}"
            errors.append(msg)
            print(f"[root_fix] ERROR: {msg}")
            continue

    print(f"[root_fix] Complete. Files updated: {changed_files}")
    if errors:
        print(f"[root_fix] Completed with errors: {len(errors)}")
    return {"changed_files": changed_files, "errors": errors}


def main():
    run_root_workbook_fixes()


if __name__ == "__main__":
    main()
